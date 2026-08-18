import argparse
import csv
import json
import random
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

# Ensure project root is in sys.path when running script directly
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from math_engine.guardrails import run_input_guardrails_suite

SEED = 20260816
COMPANY = "AsterNova Technologies Ltd."
INDUSTRY = "B2B Software & Cloud Services"
CURRENCY = "INR"
UNIT = "₹ million"
CURRENT_YEAR = "FY2025-26"
PRIOR_YEAR = "FY2024-25"
CURRENT_DATE = "2026-03-31"
PRIOR_DATE = "2025-03-31"
DEFAULT_OUTPUT_DIR = Path("Data/True_data")
MAX_ATTEMPTS = 50000

R = {
    "revenue_growth": (-0.30, 0.50),
    "gross_margin": (0.10, 0.90),
    "opex_ratio": (0.15, 0.80),
    "tax_rate": (0.15, 0.35),
    "cash_months_min": 1.0,
    "current_ratio": (1.00, 3.00),
    "dso": (30.0, 80.0),
    "dpo": (30.0, 75.0),
    "de": (0.10, 4.00),
    "capex_ratio": (0.02, 0.15),
    "dividend_payout": (0.00, 0.80),
    "allowance_pct": (0.01, 0.05),
    "it_life": (3, 7),
    "building_life": (15, 39),
}

def q(x):
    return round(float(x), 2)

def almost(a, b, tol=0.01):
    return abs(float(a) - float(b)) <= tol

def write_csv(path, rows, headers):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)

def allocate_aging(gross_ar, rng):
    shares = [
        rng.uniform(0.45, 0.60),
        rng.uniform(0.18, 0.25),
        rng.uniform(0.10, 0.16),
        rng.uniform(0.06, 0.10),
    ]
    s = sum(shares)
    shares = [x / s for x in shares]
    vals = [gross_ar * x for x in shares]
    return {
        "current": q(vals[0]),
        "days_1_30": q(vals[1]),
        "days_31_60": q(vals[2]),
        "days_61_90": q(vals[3]),
        "days_over_90": q(gross_ar - sum(vals)),
    }

def debt_maturity(total_debt, current_portion, rng):
    remaining = max(total_debt - current_portion, 0.0)
    weights = [rng.uniform(0.10, 0.25) for _ in range(4)]
    s = sum(weights)
    buckets = [remaining * w / 1.0 for w in weights]
    thereafter = max(remaining - sum(buckets), 0.0)
    return {
        "year_1": q(current_portion),
        "year_2": q(buckets[0]),
        "year_3": q(buckets[1]),
        "year_4": q(buckets[2]),
        "year_5": q(buckets[3]),
        "thereafter": q(thereafter),
        "total_debt": q(total_debt),
    }

def build_trial_balance(bs, inc, period_type, period_end, source_name, beginning_re=None, dividends=0.0):
    rows = []
    def add(code, name, fsli, acct_type, balance):
        bal = q(balance)
        rows.append({
            "account_code": code,
            "account_name": name,
            "standardized_fsli": fsli,
            "account_type": acct_type,
            "debit_amount": q(max(bal, 0.0)),
            "credit_amount": q(max(-bal, 0.0)),
            "ending_balance": bal,
        })

    add("1000", "Cash & Cash Equivalents", "Cash & Cash Equivalents", "ASSET", bs["cash_and_cash_equivalents"])
    add("1100", "Accounts Receivable", "Accounts Receivable", "ASSET", bs["accounts_receivable_gross"])
    add("1110", "Allowance for Credit Losses", "Allowance for Credit Losses", "ASSET", -bs["allowance_for_credit_losses"])
    add("1200", "Inventory", "Inventory", "ASSET", bs["inventory"])
    add("1300", "Prepaid Expenses & Other CA", "Prepaid Expenses & Other CA", "ASSET", bs["prepaid_expenses"])
    add("1500", "Property, Plant & Equipment - Gross", "Property, Plant & Equipment - Gross", "ASSET", bs["ppe_gross"])
    add("1510", "Accumulated Depreciation", "Accumulated Depreciation", "ASSET", -bs["accumulated_depreciation"])
    add("1600", "Intangible Assets", "Intangible Assets", "ASSET", bs["intangible_assets"])
    add("2000", "Accounts Payable", "Accounts Payable", "LIABILITY", -bs["accounts_payable"])
    add("2100", "Accrued Liabilities", "Accrued Liabilities", "LIABILITY", -bs["accrued_expenses"])
    add("2200", "Short-term Debt", "Short-term Debt", "LIABILITY", -bs["short_term_debt"])
    add("2210", "Current Portion of Long-Term Debt", "Current Portion of Long-Term Debt", "LIABILITY", -bs["current_portion_of_lt_debt"])
    add("2500", "Long-term Debt", "Long-term Debt", "LIABILITY", -bs["long_term_debt"])
    add("3000", "Common Stock", "Common Stock", "EQUITY", -bs["common_stock"])
    add("3100", "Additional Paid-in Capital", "Additional Paid-in Capital", "EQUITY", -bs["additional_paid_in_capital"])

    if period_type == "CURRENT_PRELIMINARY":
        add("3200", "Retained Earnings - Beginning", "Retained Earnings", "EQUITY", -beginning_re)
        add("4000", "Revenue", "Revenue", "REVENUE", -inc["revenue"])
        add("5000", "Cost of Goods Sold", "Cost of Goods Sold", "EXPENSE", inc["cogs"])
        add("5100", "Selling, General & Administrative", "Selling, General & Administrative", "EXPENSE", inc["sga_expense"])
        add("5200", "Research & Development", "Research & Development", "EXPENSE", inc["rd_expense"])
        add("5300", "Depreciation & Amortization Expense", "Depreciation & Amortization Expense", "EXPENSE", inc["depreciation_amortization"])
        add("5400", "Interest Expense", "Interest Expense", "EXPENSE", inc["interest_expense"])
        add("5500", "Income Tax Expense", "Income Tax Expense", "EXPENSE", inc["income_tax_expense"])
        if inc.get("non_operating_income", 0.0) != 0.0:
            add("5700", "Other Non-Operating Income/Expense", "Other Non-Operating Income/Expense", "REVENUE" if inc["non_operating_income"] > 0 else "EXPENSE", -inc["non_operating_income"])
        if dividends > 0:
            add("5600", "Dividends Declared", "Retained Earnings", "EQUITY", dividends)
    else:
        add("3200", "Retained Earnings", "Retained Earnings", "EQUITY", -bs["retained_earnings"])

    debit_total = q(sum(x["debit_amount"] for x in rows))
    credit_total = q(sum(x["credit_amount"] for x in rows))
    if not almost(debit_total, credit_total):
        raise ValueError(f"Trial balance unbalanced: debit={debit_total}, credit={credit_total}")

    return {
        "metadata": {
            "entity_name": COMPANY,
            "period_type": period_type,
            "period_end_date": period_end,
            "currency": CURRENCY,
            "scale": "MILLIONS",
            "source_file_name": source_name,
        },
        "accounts": rows,
    }

def make_candidate(rng):
    # PRIOR PERIOD
    prior_revenue = rng.uniform(1000.0, 1500.0)
    prior_gm = rng.uniform(0.50, 0.70)
    prior_cogs = prior_revenue * (1.0 - prior_gm)
    prior_dso = rng.uniform(40.0, 70.0)
    prior_ar_gross = prior_revenue * prior_dso / 365.0
    prior_allow = prior_ar_gross * rng.uniform(*R["allowance_pct"])
    prior_ar_net = prior_ar_gross - prior_allow
    prior_inventory = prior_revenue * rng.uniform(0.06, 0.10)
    prior_prepaid = prior_revenue * rng.uniform(0.02, 0.04)

    prior_ppe_gross = prior_revenue * rng.uniform(0.25, 0.40)
    prior_it_share = rng.uniform(0.20, 0.40)
    prior_it_life = rng.randint(R["it_life"][0], R["it_life"][1])
    prior_building_life = rng.randint(R["building_life"][0], R["building_life"][1])
    prior_it_gross = prior_ppe_gross * prior_it_share
    prior_building_gross = prior_ppe_gross - prior_it_gross
    prior_da = prior_it_gross / prior_it_life + prior_building_gross / prior_building_life
    prior_accum = prior_ppe_gross * rng.uniform(0.20, 0.35)
    prior_ppe_net = prior_ppe_gross - prior_accum
    if prior_ppe_net <= 0: raise ValueError("prior PPE net negative")
    prior_intangibles = prior_revenue * rng.uniform(0.04, 0.08)

    prior_opex_ratio = rng.uniform(0.30, 0.48)
    prior_opex = prior_revenue * prior_opex_ratio
    prior_sga_rd = prior_opex - prior_da
    if prior_sga_rd <= 0: raise ValueError("prior SG&A/R&D pool non-positive")
    prior_rd = prior_sga_rd * rng.uniform(0.18, 0.35)
    prior_sga = prior_sga_rd - prior_rd
    prior_gp = prior_revenue - prior_cogs
    prior_oi = prior_gp - prior_opex
    prior_interest = prior_revenue * rng.uniform(0.008, 0.020)
    prior_nonop = prior_revenue * rng.uniform(-0.002, 0.004)
    prior_ebt = prior_oi - prior_interest + prior_nonop
    if prior_ebt <= 0: raise ValueError("prior EBT non-positive")
    prior_tax = prior_ebt * rng.uniform(0.20, 0.30)
    prior_ni = prior_ebt - prior_tax

    common_stock = rng.uniform(90.0, 160.0)
    apic = rng.uniform(100.0, 190.0)
    prior_re = rng.uniform(130.0, 280.0)
    prior_equity = common_stock + apic + prior_re

    prior_current_liabilities = prior_revenue * rng.uniform(0.18, 0.30)
    prior_st_debt = prior_current_liabilities * rng.uniform(0.05, 0.15)
    prior_current_ltd = prior_current_liabilities * rng.uniform(0.10, 0.25)
    prior_ap = cogs * (rng.uniform(40.0, 70.0) / 365.0)
    prior_accrued = prior_current_liabilities - prior_st_debt - prior_current_ltd - prior_ap
    if prior_accrued <= 0: raise ValueError("prior accrued negative")
    prior_ltd = prior_revenue * rng.uniform(0.12, 0.25)
    prior_liabilities = prior_current_liabilities + prior_ltd

    prior_cash = prior_liabilities + prior_equity - (prior_ar_net + prior_inventory + prior_prepaid + prior_ppe_net + prior_intangibles)
    if prior_cash <= 0: raise ValueError("prior cash negative")

    prior_bs = {
        "cash_and_cash_equivalents": q(prior_cash),
        "accounts_receivable_gross": q(prior_ar_gross),
        "allowance_for_credit_losses": q(prior_allow),
        "accounts_receivable_net": q(prior_ar_net),
        "inventory": q(prior_inventory),
        "prepaid_expenses": q(prior_prepaid),
        "total_current_assets": q(prior_cash + prior_ar_net + prior_inventory + prior_prepaid),
        "ppe_gross": q(prior_ppe_gross),
        "accumulated_depreciation": q(prior_accum),
        "ppe_net": q(prior_ppe_net),
        "intangible_assets": q(prior_intangibles),
        "total_non_current_assets": q(prior_ppe_net + prior_intangibles),
        "total_assets": q(prior_cash + prior_ar_net + prior_inventory + prior_prepaid + prior_ppe_net + prior_intangibles),
        "accounts_payable": q(prior_ap),
        "accrued_expenses": q(prior_accrued),
        "short_term_debt": q(prior_st_debt),
        "current_portion_of_lt_debt": q(prior_current_ltd),
        "total_current_liabilities": q(prior_current_liabilities),
        "long_term_debt": q(prior_ltd),
        "total_non_current_liabilities": q(prior_ltd),
        "total_liabilities": q(prior_liabilities),
        "common_stock": q(common_stock),
        "additional_paid_in_capital": q(apic),
        "retained_earnings": q(prior_re),
        "aoci": 0.0,
        "treasury_stock": 0.0,
        "total_equity": q(prior_equity),
    }
    prior_inc = {
        "revenue": q(prior_revenue),
        "cogs": q(prior_cogs),
        "gross_profit": q(prior_gp),
        "sga_expense": q(prior_sga),
        "rd_expense": q(prior_rd),
        "depreciation_amortization": q(prior_da),
        "total_operating_expenses": q(prior_opex),
        "operating_income": q(prior_oi),
        "interest_expense": q(prior_interest),
        "non_operating_income": q(prior_nonop),
        "income_tax_expense": q(prior_tax),
        "net_income": q(prior_ni),
    }

    # CURRENT PERIOD
    growth = rng.uniform(R["revenue_growth"][0] + 0.05, R["revenue_growth"][1] - 0.05)
    revenue = prior_revenue * (1.0 + growth)
    gross_margin = rng.uniform(0.50, 0.70)
    cogs = revenue * (1.0 - gross_margin)
    gross_profit = revenue - cogs

    dso = rng.uniform(R["dso"][0] + 5, R["dso"][1] - 5)
    ar_gross = revenue * dso / 365.0
    allowance = ar_gross * rng.uniform(*R["allowance_pct"])
    ar_net = ar_gross - allowance
    inventory = revenue * rng.uniform(0.06, 0.10)
    prepaid = revenue * rng.uniform(0.02, 0.04)

    capex = revenue * rng.uniform(R["capex_ratio"][0] + 0.01, R["capex_ratio"][1] - 0.01)
    current_gross_ppe = prior_ppe_gross + capex
    it_share = rng.uniform(0.20, 0.40)
    it_life = rng.randint(R["it_life"][0], R["it_life"][1])
    building_life = rng.randint(R["building_life"][0], R["building_life"][1])
    it_gross = current_gross_ppe * it_share
    building_gross = current_gross_ppe - it_gross
    da = max(it_gross / it_life + building_gross / building_life, prior_da * 1.05)
    current_accum = prior_accum + da
    ppe_net = current_gross_ppe - current_accum
    if ppe_net <= 0: raise ValueError("current PPE net negative")
    intangible_assets = prior_intangibles * rng.uniform(0.98, 1.05)

    total_opex = max(revenue * rng.uniform(0.30, 0.48), da + 10.0)
    sga_rd = total_opex - da
    if sga_rd <= 0: raise ValueError("current SG&A/R&D pool non-positive")
    rd = sga_rd * rng.uniform(0.18, 0.35)
    sga = sga_rd - rd
    operating_income = gross_profit - total_opex
    interest = revenue * rng.uniform(0.008, 0.020)
    nonop = revenue * rng.uniform(-0.002, 0.004)
    ebt = operating_income - interest + nonop
    if ebt <= 0: raise ValueError("current EBT non-positive")
    tax = ebt * rng.uniform(0.18, 0.30)
    net_income = ebt - tax

    dividends = net_income * rng.uniform(R["dividend_payout"][0] + 0.05, R["dividend_payout"][1] - 0.05)
    ending_re = prior_re + net_income - dividends
    if ending_re <= 0: raise ValueError("ending RE negative")
    equity = common_stock + apic + ending_re

    current_liabilities = revenue * rng.uniform(0.18, 0.30)
    short_term_debt = current_liabilities * rng.uniform(0.05, 0.15)
    current_ltd = current_liabilities * rng.uniform(0.10, 0.25)
    dpo = rng.uniform(R["dpo"][0], R["dpo"][1])
    accounts_payable = cogs * (dpo / 365.0)
    accrued = current_liabilities - short_term_debt - current_ltd - accounts_payable
    if accrued <= 0: raise ValueError("current accrued negative")
    long_term_debt = revenue * rng.uniform(0.12, 0.25)
    total_liabilities = current_liabilities + long_term_debt

    cash = total_liabilities + equity - (ar_net + inventory + prepaid + ppe_net + intangible_assets)
    if cash <= 0: raise ValueError("current cash negative")

    current_bs = {
        "cash_and_cash_equivalents": q(cash),
        "accounts_receivable_gross": q(ar_gross),
        "allowance_for_credit_losses": q(allowance),
        "accounts_receivable_net": q(ar_net),
        "inventory": q(inventory),
        "prepaid_expenses": q(prepaid),
        "total_current_assets": q(cash + ar_net + inventory + prepaid),
        "ppe_gross": q(current_gross_ppe),
        "accumulated_depreciation": q(current_accum),
        "ppe_net": q(ppe_net),
        "intangible_assets": q(intangible_assets),
        "total_non_current_assets": q(ppe_net + intangible_assets),
        "total_assets": q(cash + ar_net + inventory + prepaid + ppe_net + intangible_assets),
        "accounts_payable": q(accounts_payable),
        "accrued_expenses": q(accrued),
        "short_term_debt": q(short_term_debt),
        "current_portion_of_lt_debt": q(current_ltd),
        "total_current_liabilities": q(current_liabilities),
        "long_term_debt": q(long_term_debt),
        "total_non_current_liabilities": q(long_term_debt),
        "total_liabilities": q(total_liabilities),
        "common_stock": q(common_stock),
        "additional_paid_in_capital": q(apic),
        "retained_earnings": q(ending_re),
        "aoci": 0.0,
        "treasury_stock": 0.0,
        "total_equity": q(equity),
    }
    current_inc = {
        "revenue": q(revenue),
        "cogs": q(cogs),
        "gross_profit": q(gross_profit),
        "sga_expense": q(sga),
        "rd_expense": q(rd),
        "depreciation_amortization": q(da),
        "total_operating_expenses": q(total_opex),
        "operating_income": q(operating_income),
        "interest_expense": q(interest),
        "non_operating_income": q(nonop),
        "income_tax_expense": q(tax),
        "net_income": q(net_income),
    }

    debt_change = (short_term_debt + current_ltd + long_term_debt) - (prior_st_debt + prior_current_ltd + prior_ltd)
    financing_cf = debt_change - dividends
    investing_cf = -capex
    target_net_cash_change = cash - prior_cash
    operating_cf = target_net_cash_change - investing_cf - financing_cf
    working_capital_changes = operating_cf - net_income - da

    current_cf = {
        "net_income_starting": q(net_income),
        "depreciation_addback": q(da),
        "working_capital_changes": q(working_capital_changes),
        "operating_cash_flow": q(operating_cf),
        "capital_expenditures": q(-capex),
        "investing_cash_flow": q(investing_cf),
        "debt_repayments_or_borrowings": q(debt_change),
        "dividends_paid": q(-dividends),
        "financing_cash_flow": q(financing_cf),
        "net_cash_change": q(target_net_cash_change),
        "beginning_cash": q(prior_cash),
        "ending_cash": q(cash),
    }

    ar_age = allocate_aging(ar_gross, rng)
    ar_age.update({"gross_ar": q(ar_gross), "allowance_for_credit_losses": q(allowance), "net_ar": q(ar_net)})

    ppe_schedule = {
        "gross_ppe": q(current_gross_ppe),
        "accumulated_depreciation": q(current_accum),
        "net_ppe": q(ppe_net),
        "additions_capex": q(capex),
        "disposals": 0.0,
        "depreciation_expense": q(da),
        "asset_classes": [
            {"asset_class": "IT hardware/software", "gross_balance": q(it_gross), "useful_life_years": it_life},
            {"asset_class": "Buildings / real property", "gross_balance": q(building_gross), "useful_life_years": building_life},
        ],
    }
    total_debt = short_term_debt + current_ltd + long_term_debt
    debt_schedule = debt_maturity(total_debt, current_ltd, rng)

    return {
        "prior": {
            "balance_sheet": prior_bs,
            "income_statement": prior_inc,
            "bs": prior_bs,
            "inc": prior_inc,
        },
        "current": {
            "balance_sheet": current_bs,
            "income_statement": current_inc,
            "cash_flow_statement": current_cf,
            "equity_statement": {
                "beginning_retained_earnings": q(prior_re),
                "net_income": q(net_income),
                "dividends_declared": q(dividends),
                "ending_retained_earnings": q(ending_re),
            },
            "footnotes": {
                "ar_aging": ar_age,
                "ppe_sched": ppe_schedule,
                "debt_maturity": debt_schedule,
            },
            "bs": current_bs,
            "inc": current_inc,
            "cf": current_cf,
        },
        "equity": {
            "beginning_retained_earnings": q(prior_re),
            "net_income": q(net_income),
            "dividends_declared": q(dividends),
            "ending_retained_earnings": q(ending_re),
        },
        "footnotes": {
            "ar_aging": ar_age,
            "ppe_sched": ppe_schedule,
            "debt_maturity": debt_schedule,
        }
    }

def guardrails(d):
    c = d["current"]
    p = d["prior"]
    return run_input_guardrails_suite(current_data=c, prior_data=p)

# ----------------------------
# Excel & PDF Rendering Helpers
# ----------------------------

def excel_header(ws, title, subtitle):
    ws["A1"] = COMPANY
    ws["A2"] = title
    ws["A3"] = subtitle
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"].font = Font(italic=True, size=10)

def write_table_xlsx(path, title, subtitle, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    excel_header(ws, title, subtitle)
    fill = PatternFill("solid", fgColor="D9EAF7")
    side = Side(style="thin", color="B7B7B7")
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = Font(bold=True)
        c.fill = fill
        c.border = Border(bottom=side)
    for r_idx, row_vals in enumerate(rows, 6):
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 35
    wb.save(path)

def write_trial_balance_xlsx(path, title, subtitle, sheet_name, tb_data):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    excel_header(ws, title, subtitle)
    fill = PatternFill("solid", fgColor="D9EAF7")
    side = Side(style="thin", color="B7B7B7")
    headers = ["Account Code", "Account Name", "Standardized FSLI", "Account Type", f"Debit ({UNIT})", f"Credit ({UNIT})", f"Ending Balance ({UNIT})"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(5, col, h)
        c.font = Font(bold=True)
        c.fill = fill
        c.border = Border(bottom=side)
    for r_idx, acct in enumerate(tb_data["accounts"], 6):
        row_vals = [
            acct["account_code"],
            acct["account_name"],
            acct["standardized_fsli"],
            acct["account_type"],
            acct["debit_amount"],
            acct["credit_amount"],
            acct["ending_balance"]
        ]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(r_idx, c_idx, val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 28
    wb.save(path)

def make_single_pdf(path, title, subtitle, headers, rows):
    styles = getSampleStyleSheet()
    story = [
        Paragraph(COMPANY, styles["Title"]),
        Paragraph(title, styles["Heading2"]),
        Paragraph(subtitle, styles["Normal"]),
        Spacer(1, 12)
    ]
    data = [headers] + [[str(f"{x:,.2f}" if isinstance(x, (int, float)) else x) for x in r] for r in rows]
    t = Table(data, colWidths=[340, 150], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9)
    ]))
    story.append(t)
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    doc.build(story)

# ----------------------------
# Main Pipeline
# ----------------------------

def main():
    parser = argparse.ArgumentParser(description="Clean Financial Statement Data Generator")
    parser.add_argument("--format", choices=["excel", "pdf"], default="excel", help="Output format for financial statements and footnotes")
    parser.add_argument("--output-dir", type=str, default=None, help="Target output directory (default: Data/True_data)")
    args = parser.parse_args()

    output_format = args.format
    output_dir = Path(args.output_dir) if args.output_dir else DEFAULT_OUTPUT_DIR

    print(f"Generating clean financial dataset [format={output_format.upper()}] -> {output_dir}...")

    rng = random.Random(SEED)
    if output_dir.exists():
        shutil.rmtree(output_dir)
    
    prior_dir = output_dir / "prior_data"
    curr_dir = output_dir / "current_data"
    footnotes_dir = curr_dir / "footnotes"
    footnotes_dir.mkdir(parents=True)
    prior_dir.mkdir(parents=True)

    selected = None
    for _ in range(MAX_ATTEMPTS):
        try:
            candidate = make_candidate(rng)
            candidate["current"]["max_contingent_liability_exposure"] = q(candidate["current"]["bs"]["cash_and_cash_equivalents"] * rng.uniform(0.10, 0.40))
            candidate["contingent_liability"] = {"maximum_exposure": candidate["current"]["max_contingent_liability_exposure"]}
            
            current_tb = build_trial_balance(candidate["current"]["bs"], candidate["current"]["inc"], "CURRENT_PRELIMINARY", CURRENT_DATE, "preliminary_trial_balance.xlsx", candidate["equity"]["beginning_retained_earnings"], candidate["equity"]["dividends_declared"])
            prior_tb = build_trial_balance(candidate["prior"]["bs"], candidate["prior"]["inc"], "PRIOR_AUDITED", PRIOR_DATE, "final_trial_balance.xlsx")
            
            candidate["current"]["preliminary_trial_balance"] = current_tb
            candidate["prior"]["final_trial_balance"] = prior_tb

            gr = guardrails(candidate)
            if all(x["status"] == "PASS" for x in gr):
                selected = (candidate, current_tb, prior_tb, gr)
                break
        except Exception:
            continue

    if selected is None:
        raise RuntimeError("Failed to generate valid dataset satisfying all sanity guardrails.")

    d, current_tb, prior_tb, gr = selected
    c = d["current"]
    p = d["prior"]
    fn = d["footnotes"]

    # Save Trial Balances
    write_trial_balance_xlsx(
        prior_dir / "final_trial_balance.xlsx",
        "Audited Final Trial Balance — Prior",
        f"{PRIOR_YEAR} | Amounts in {UNIT}",
        "Final Trial Balance",
        prior_tb
    )

    write_trial_balance_xlsx(
        curr_dir / "preliminary_trial_balance.xlsx",
        "Preliminary Trial Balance — Current",
        f"{CURRENT_YEAR} | Amounts in {UNIT}",
        "Preliminary Trial Balance",
        current_tb
    )

    def save_statement(path_no_ext, title, subtitle, sheet_name, headers, rows):
        if output_format == "excel":
            write_table_xlsx(path_no_ext.with_suffix(".xlsx"), title, subtitle, sheet_name, headers, rows)
        else:
            make_single_pdf(path_no_ext.with_suffix(".pdf"), title, subtitle, headers, rows)

    # Save Prior Data Statements
    pbs = p["bs"]
    pbs_rows = [
        ("Cash & Cash Equivalents", pbs["cash_and_cash_equivalents"]),
        ("Accounts Receivable, net", pbs["accounts_receivable_net"]),
        ("Inventory", pbs["inventory"]),
        ("Prepaid Expenses & Other Current Assets", pbs["prepaid_expenses"]),
        ("Total Current Assets", pbs["total_current_assets"]),
        ("Property, Plant & Equipment, net", pbs["ppe_net"]),
        ("Intangible Assets", pbs["intangible_assets"]),
        ("Total Assets", pbs["total_assets"]),
        ("Accounts Payable", pbs["accounts_payable"]),
        ("Accrued Liabilities", pbs["accrued_expenses"]),
        ("Short-term Debt", pbs["short_term_debt"]),
        ("Current Portion of Long-Term Debt", pbs["current_portion_of_lt_debt"]),
        ("Total Current Liabilities", pbs["total_current_liabilities"]),
        ("Long-term Debt", pbs["long_term_debt"]),
        ("Total Liabilities", pbs["total_liabilities"]),
        ("Common Stock", pbs["common_stock"]),
        ("Additional Paid-in Capital", pbs["additional_paid_in_capital"]),
        ("Retained Earnings", pbs["retained_earnings"]),
        ("Total Equity", pbs["total_equity"]),
        ("Total Liabilities & Equity", pbs["total_liabilities"] + pbs["total_equity"])
    ]
    save_statement(prior_dir / "balance_sheet", "Audited Balance Sheet — Prior", f"{PRIOR_YEAR} | Amounts in {UNIT}", "Balance Sheet", ["Line Item", UNIT], pbs_rows)

    pinc = p["inc"]
    pinc_rows = [
        ("Revenue", pinc["revenue"]),
        ("Cost of Goods Sold", pinc["cogs"]),
        ("Gross Profit", pinc["gross_profit"]),
        ("Selling, General & Administrative", pinc["sga_expense"]),
        ("Research & Development", pinc["rd_expense"]),
        ("Depreciation & Amortization", pinc["depreciation_amortization"]),
        ("Total Operating Expenses", pinc["total_operating_expenses"]),
        ("Operating Income", pinc["operating_income"]),
        ("Interest Expense", pinc["interest_expense"]),
        ("Non-Operating Income / (Expense)", pinc["non_operating_income"]),
        ("Income Before Tax", pinc["operating_income"] + pinc["non_operating_income"] - pinc["interest_expense"]),
        ("Income Tax Expense", pinc["income_tax_expense"]),
        ("Net Income", pinc["net_income"])
    ]
    save_statement(prior_dir / "income_statement", "Audited Income Statement — Prior", f"{PRIOR_YEAR} | Amounts in {UNIT}", "Income Statement", ["Line Item", UNIT], pinc_rows)

    # Save Current Data Statements
    cbs = c["bs"]
    cbs_rows = [
        ("Cash & Cash Equivalents", cbs["cash_and_cash_equivalents"]),
        ("Accounts Receivable, net", cbs["accounts_receivable_net"]),
        ("Inventory", cbs["inventory"]),
        ("Prepaid Expenses & Other Current Assets", cbs["prepaid_expenses"]),
        ("Total Current Assets", cbs["total_current_assets"]),
        ("Property, Plant & Equipment, net", cbs["ppe_net"]),
        ("Intangible Assets", cbs["intangible_assets"]),
        ("Total Assets", cbs["total_assets"]),
        ("Accounts Payable", cbs["accounts_payable"]),
        ("Accrued Liabilities", cbs["accrued_expenses"]),
        ("Short-term Debt", cbs["short_term_debt"]),
        ("Current Portion of Long-Term Debt", cbs["current_portion_of_lt_debt"]),
        ("Total Current Liabilities", cbs["total_current_liabilities"]),
        ("Long-term Debt", cbs["long_term_debt"]),
        ("Total Liabilities", cbs["total_liabilities"]),
        ("Common Stock", cbs["common_stock"]),
        ("Additional Paid-in Capital", cbs["additional_paid_in_capital"]),
        ("Retained Earnings", cbs["retained_earnings"]),
        ("Total Equity", cbs["total_equity"]),
        ("Total Liabilities & Equity", cbs["total_liabilities"] + cbs["total_equity"])
    ]
    save_statement(curr_dir / "balance_sheet", "Preliminary Balance Sheet — Current", f"{CURRENT_YEAR} | Amounts in {UNIT}", "Balance Sheet", ["Line Item", UNIT], cbs_rows)

    cinc = c["inc"]
    cinc_rows = [
        ("Revenue", cinc["revenue"]),
        ("Cost of Goods Sold", cinc["cogs"]),
        ("Gross Profit", cinc["gross_profit"]),
        ("Selling, General & Administrative", cinc["sga_expense"]),
        ("Research & Development", cinc["rd_expense"]),
        ("Depreciation & Amortization", cinc["depreciation_amortization"]),
        ("Total Operating Expenses", cinc["total_operating_expenses"]),
        ("Operating Income", cinc["operating_income"]),
        ("Interest Expense", cinc["interest_expense"]),
        ("Non-Operating Income / (Expense)", cinc["non_operating_income"]),
        ("Income Before Tax", cinc["operating_income"] + cinc["non_operating_income"] - cinc["interest_expense"]),
        ("Income Tax Expense", cinc["income_tax_expense"]),
        ("Net Income", cinc["net_income"])
    ]
    save_statement(curr_dir / "income_statement", "Preliminary Income Statement — Current", f"{CURRENT_YEAR} | Amounts in {UNIT}", "Income Statement", ["Line Item", UNIT], cinc_rows)

    ccf = c["cf"]
    ccf_rows = [
        ("Net Income", ccf["net_income_starting"]),
        ("Depreciation & Amortization Add-back", ccf["depreciation_addback"]),
        ("Working Capital Changes", ccf["working_capital_changes"]),
        ("Operating Cash Flow", ccf["operating_cash_flow"]),
        ("Capital Expenditures", ccf["capital_expenditures"]),
        ("Investing Cash Flow", ccf["investing_cash_flow"]),
        ("Debt Borrowings / (Repayments)", ccf["debt_repayments_or_borrowings"]),
        ("Dividends Paid", ccf["dividends_paid"]),
        ("Financing Cash Flow", ccf["financing_cash_flow"]),
        ("Net Cash Change", ccf["net_cash_change"]),
        ("Beginning Cash", ccf["beginning_cash"]),
        ("Ending Cash", ccf["ending_cash"])
    ]
    save_statement(curr_dir / "cash_flow_statement", "Preliminary Cash Flow Statement — Current", f"{CURRENT_YEAR} | Amounts in {UNIT}", "Cash Flow", ["Line Item", UNIT], ccf_rows)

    ceq = d["equity"]
    ceq_rows = [
        ("Beginning Retained Earnings", ceq["beginning_retained_earnings"]),
        ("Net Income", ceq["net_income"]),
        ("Dividends Declared", -ceq["dividends_declared"]),
        ("Ending Retained Earnings", ceq["ending_retained_earnings"])
    ]
    save_statement(curr_dir / "equity_statement", "Statement of Stockholders' Equity", f"{CURRENT_YEAR} | Amounts in {UNIT}", "Stockholders Equity", ["Line Item", UNIT], ceq_rows)

    # Save Footnotes
    arag = fn["ar_aging"]
    ar_rows = [
        ("Current", arag["current"]),
        ("1-30 Days", arag["days_1_30"]),
        ("31-60 Days", arag["days_31_60"]),
        ("61-90 Days", arag["days_61_90"]),
        ("Over 90 Days", arag["days_over_90"]),
        ("Gross Accounts Receivable", arag["gross_ar"]),
        ("Allowance for Credit Losses", -arag["allowance_for_credit_losses"]),
        ("Net Accounts Receivable", arag["net_ar"])
    ]
    save_statement(footnotes_dir / "ar_aging", "Footnote: Accounts Receivable Aging", f"{CURRENT_YEAR} | Amounts in {UNIT}", "AR Aging", ["Bucket", UNIT], ar_rows)

    ppe = fn["ppe_sched"]
    ppe_rows = [
        ("Gross PP&E", ppe["gross_ppe"]),
        ("Additions / CapEx", ppe["additions_capex"]),
        ("Disposals", -ppe["disposals"]),
        ("Depreciation Expense", -ppe["depreciation_expense"]),
        ("Accumulated Depreciation", -ppe["accumulated_depreciation"]),
        ("Net PP&E", ppe["net_ppe"])
    ]
    save_statement(footnotes_dir / "ppe_sched", "Footnote: PP&E Schedule", f"{CURRENT_YEAR} | Amounts in {UNIT}", "PPE Schedule", ["Line Item", UNIT], ppe_rows)

    dm = fn["debt_maturity"]
    dm_rows = [
        ("Year 1", dm["year_1"]),
        ("Year 2", dm["year_2"]),
        ("Year 3", dm["year_3"]),
        ("Year 4", dm["year_4"]),
        ("Year 5", dm["year_5"]),
        ("Thereafter", dm["thereafter"]),
        ("Total Debt", dm["total_debt"])
    ]
    save_statement(footnotes_dir / "debt_maturity", "Footnote: Debt Maturity Schedule", f"{CURRENT_YEAR} | Amounts in {UNIT}", "Debt Maturity", ["Maturity", UNIT], dm_rows)

    # Save Guardrail Results CSV Summary
    write_csv(output_dir / "guardrail_results.csv", gr, ["rule_id", "category", "rule_name", "status", "value", "benchmark", "message"])
    (output_dir / "injected_flaws_ground_truth.json").write_text(json.dumps([], indent=2), encoding="utf-8")

    print(f"Clean dataset successfully generated in {output_format.upper()} format into: {output_dir.resolve()}")

if __name__ == "__main__":
    main()
