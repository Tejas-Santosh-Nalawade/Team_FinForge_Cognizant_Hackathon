from __future__ import annotations

import csv
import json
import re
import zipfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable, Optional
from xml.etree import ElementTree as ET

from backend.app.core.normalization.coa_mapper import CanonicalMapper
from backend.app.core.parser.spell_checker import lint_text
from backend.app.core.normalization.source_router import route_source
from backend.app.core.normalization.normalization_utils import infer_unit_metadata, convert_scale, normalize_sign

SUPPORTED = {".xlsx", ".xlsm", ".csv", ".json", ".jsonl", ".pdf", ".docx", ".txt", ".md"}
FINANCIAL_CONTEXTS = {
    "balance_sheet", "income_statement", "cash_flow_statement", "equity_statement",
    "ar_aging", "ppe_sched", "debt_maturity",
}
REQUIRED = {
    "balance_sheet": {"total_assets", "total_liabilities", "total_equity"},
    "income_statement": {"revenue", "cogs", "gross_profit", "total_operating_expenses", "operating_income", "net_income"},
    "cash_flow_statement": {"net_income_starting", "operating_cash_flow", "investing_cash_flow", "financing_cash_flow", "net_cash_change", "ending_cash"},
}
IGNORED_SUPPLEMENTARY = {
    "total liabilities equity", "total liabilities and equity", "income before tax", "earnings before tax", "ebt",
}
BUDGET_CATEGORY = {
    "REVENUE": "Revenue", "COGS": "Direct Cost", "SGA": "Operating Expense",
    "RND": "Operating Expense", "DA": "Operating Expense", "OPEX": "Operating Expense",
    "CAPEX": "Capital Expenditure", "TAX_EXPENSE": "Tax", "NET_INCOME": "Revenue",
}
DRIVER_MAP = {
    "headcount": ("HEADCOUNT", "HEADCOUNT"),
    "employee count": ("HEADCOUNT", "HEADCOUNT"),
    "fte": ("HEADCOUNT", "HEADCOUNT"),
    "operating volume": ("VOLUME", "OPERATING_VOLUME"),
    "sales volume": ("VOLUME", "OPERATING_VOLUME"),
    "production volume": ("VOLUME", "OPERATING_VOLUME"),
    "price per unit": ("PRICING", "PRICE_PER_UNIT"),
    "average selling price": ("PRICING", "PRICE_PER_UNIT"),
    "capacity": ("CAPACITY", None),
    "churn": ("CHURN", None),
    "customer churn": ("CHURN", None),
}
FORECAST_MAP = {
    "revenue": "revenue", "sales": "revenue", "turnover": "revenue",
    "cogs": "cogs", "cost of goods sold": "cogs", "cost of sales": "cogs", "cost of revenue": "cogs",
    "opex": "opex", "operating expenses": "opex", "total operating expenses": "opex",
    "operating income": "operating_income", "ebit": "operating_income",
    "capex": "capex", "capital expenditures": "capex", "capital expenditure": "capex",
}


def norm(s: Any) -> str:
    s = str(s or "").lower().replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def parse_number(v: Any) -> Optional[float]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace(",", "").replace("₹", "").replace("$", "").replace("€", "").replace("£", "")
    s = re.sub(r"\b(INR|USD|EUR|GBP)\b", "", s, flags=re.I).strip()
    s = re.sub(r"[%]$", "", s).strip()
    try:
        x = float(s.strip("() "))
        return -x if neg else x
    except ValueError:
        return None


def discover(inputs: Iterable[str | Path]) -> list[Path]:
    out: list[Path] = []
    for x in inputs:
        p = Path(x)
        if p.is_file() and p.suffix.lower() in SUPPORTED:
            out.append(p)
        elif p.is_dir():
            out.extend(q for q in p.rglob("*") if q.is_file() and q.suffix.lower() in SUPPORTED)
    return sorted(set(out))


def infer_context(path: Path, sheet: str = "", title: str = "") -> Optional[str]:
    text = norm(" ".join([str(path), sheet, title]))
    if "trial balance" in text or "preliminary trial" in text or "final trial" in text:
        return "trial_balance"
    if "ar aging" in text or "accounts receivable aging" in text or "receivable aging" in text:
        return "ar_aging"
    if "debt maturity" in text or "maturity schedule" in text:
        return "debt_maturity"
    if "ppe sched" in text or "ppe schedule" in text or "pp e schedule" in text:
        return "ppe_sched"
    if "cash flow" in text:
        return "cash_flow_statement"
    if "income statement" in text or "profit and loss" in text or "statement of operations" in text:
        return "income_statement"
    if "balance sheet" in text or "statement of financial position" in text:
        return "balance_sheet"
    if "equity statement" in text or "stockholders equity" in text or "shareholders equity" in text or "statement of changes in equity" in text:
        return "equity_statement"
    return None


def infer_period_kind(path: Path, title: str = "") -> str:
    t = norm(f"{path} {title}")
    if any(k in t for k in ["prior data", "prior", "audited", "final trial"]):
        return "prior"
    return "current"


def infer_entity(path: Path, text: str = "") -> str | None:
    """Best-effort legal-entity detection without inventing an entity.

    Supported conventions include folders such as entities/<name>,
    subsidiaries/<name>, entity_<name>, and explicit Entity:/Company:/Subsidiary:
    headers inside a source file.
    """
    m = re.search(r"(?im)^\s*(?:entity|legal entity|subsidiary|company)\s*[:=-]\s*([^|\n]{2,120})\s*$", text or "")
    if m:
        value = m.group(1).strip()
        if value:
            return value
    parts = list(path.parts)
    markers = {"entities", "entity", "subsidiaries", "subsidiary", "companies", "company", "legal_entities", "legal-entities"}
    for i, part in enumerate(parts[:-1]):
        if norm(part).replace(" ", "_") in markers and i + 1 < len(parts) - 1:
            candidate = parts[i + 1].strip()
            if candidate:
                return candidate
    for part in reversed(parts[:-1]):
        m = re.match(r"(?i)^(?:entity|subsidiary|company)[ _-]+(.+)$", part.strip())
        if m and m.group(1).strip():
            return m.group(1).strip()
    return None


def _consolidate_trial_balances(tbs: list[dict]) -> dict:
    if not tbs:
        return {}
    if len(tbs) == 1:
        return tbs[0]
    accounts = []
    entities = []
    for tb in tbs:
        meta = tb.get("metadata") or {}
        entity = meta.get("entity_name")
        if entity and entity not in entities:
            entities.append(entity)
        for account in tb.get("accounts") or []:
            rec = dict(account)
            if entity:
                rec.setdefault("source_entity", entity)
            accounts.append(rec)
    return {
        "metadata": {
            "entity_name": "Consolidated",
            "consolidated_entities": entities,
            "period_type": (tbs[0].get("metadata") or {}).get("period_type"),
            "period_end_date": (tbs[0].get("metadata") or {}).get("period_end_date"),
            "currency": (tbs[0].get("metadata") or {}).get("currency"),
            "scale": (tbs[0].get("metadata") or {}).get("scale"),
        },
        "accounts": accounts,
    }


def _fy_period_end(raw: str) -> str | None:
    """Convert FY2025-26 / FY2025-2026 to the conventional fiscal year-end date."""
    m = re.search(r"\bFY\s*(20\d{2})\s*[-/]\s*(\d{2,4})\b", raw, re.I)
    if not m:
        return None
    start_year = int(m.group(1))
    tail = m.group(2)
    end_year = int(tail) if len(tail) == 4 else (start_year // 100) * 100 + int(tail)
    if end_year < start_year:
        end_year += 100
    return f"{end_year:04d}-03-31"


def extract_metadata_text(text: str) -> dict[str, Any]:
    hints: dict[str, Any] = {}
    lines = [x.strip() for x in text.splitlines() if x.strip()]
    if lines:
        first = lines[0]
        first_n = norm(first)
        header_like = any(k in first_n for k in [
            "balance sheet", "income statement", "cash flow", "trial balance", "footnote",
            "company period metric", "company period driver", "line item", "account code"
        ])
        if len(first) < 150 and "|" not in first and not header_like:
            hints["client_name"] = first
    lo = text.lower()
    if "₹" in text or " inr" in lo or "rupee" in lo:
        hints["currency"] = "INR"
    elif "$" in text or " usd" in lo:
        hints["currency"] = "USD"
    elif "€" in text or " eur" in lo:
        hints["currency"] = "EUR"
    elif "£" in text or " gbp" in lo:
        hints["currency"] = "GBP"
    if "million" in lo or " million" in lo:
        hints["scale"] = "MILLIONS"
    elif "billion" in lo:
        hints["scale"] = "BILLIONS"
    elif "thousand" in lo or "'000" in lo:
        hints["scale"] = "THOUSANDS"
    m = re.search(r"\bFY\s*(20\d{2})(?:\s*[-/]\s*(\d{2,4}))?\b", text, re.I)
    if m:
        hints["fiscal_year"] = f"FY{m.group(1)}"
        fy_end = _fy_period_end(m.group(0))
        if fy_end:
            hints["period"] = fy_end
    d = re.search(r"\b(20\d{2})[-/](\d{2})[-/](\d{2})\b", text)
    if d:
        hints["period"] = f"{d.group(1)}-{d.group(2)}-{d.group(3)}"
    return hints


@dataclass
class Item:
    label: str
    value: float
    context: str
    period_kind: str
    source_file: str
    source_location: str
    raw_text: str = ""
    source_currency: Optional[str] = None
    source_scale: Optional[str] = None
    entity_name: Optional[str] = None


class Layer2Gateway:
    def __init__(self, taxonomy_path: str | Path | None = None, fuzzy_threshold: float = 0.88):
        kwargs = {"fuzzy_threshold": fuzzy_threshold}
        if taxonomy_path:
            kwargs["taxonomy_path"] = taxonomy_path
        self.mapper = CanonicalMapper(**kwargs)

    def run(self, inputs: Iterable[str | Path], *, client_name: str | None = None,
            period: str | None = None, comparative_period: str | None = None,
            currency: str | None = None, scale: str | None = None,
            framework: str | None = None, include_qualitative: bool = False) -> dict:
        paths = discover(inputs)
        hints: dict[str, Any] = {}
        items: list[Item] = []
        trial_balances: dict[str, list[dict]] = {"prior": [], "current": []}
        planning_rows: list[dict] = []
        text_corpus: list[dict] = []
        parse_issues: list[dict] = []

        source_routes: list[dict] = []
        routed_paths: list[Path] = []
        for p in paths:
            route = route_source(p, include_qualitative=include_qualitative)
            source_routes.append({"source_file": str(p), **route.to_dict()})
            if route.channel == "ignored":
                continue
            routed_paths.append(p)
            try:
                if p.suffix.lower() in {".xlsx", ".xlsm"}:
                    self._parse_excel(p, items, trial_balances, planning_rows, text_corpus, hints)
                elif p.suffix.lower() == ".csv":
                    self._parse_csv(p, items, text_corpus)
                elif p.suffix.lower() == ".json":
                    self._parse_json(p, items, planning_rows, text_corpus)
                elif p.suffix.lower() == ".jsonl":
                    if route.channel == "qualitative" or include_qualitative:
                        self._parse_jsonl_text(p, text_corpus)
                elif p.suffix.lower() == ".pdf":
                    if route.channel == "qualitative" or infer_context(p):
                        self._parse_pdf(p, items, text_corpus, hints)
                elif p.suffix.lower() == ".docx":
                    if route.channel == "qualitative" or include_qualitative:
                        self._parse_docx(p, text_corpus)
                elif p.suffix.lower() in {".txt", ".md"}:
                    if route.channel == "qualitative" or include_qualitative:
                        self._parse_text(p, text_corpus)
            except Exception as e:
                parse_issues.append({"source_file": str(p), "error": f"{type(e).__name__}: {e}"})

        detected_current_period = hints.get("current_period") or hints.get("period")
        detected_prior_period = hints.get("comparative_period")
        metadata = {
            "client_name": client_name or hints.get("client_name"),
            "period": period or detected_current_period,
            "currency": currency or hints.get("currency"),
            "scale": scale or hints.get("scale"),
            "review_stage": "CY_DRAFT_FS",
            "entity_name": client_name or hints.get("client_name"),
            "period_end_date": period or detected_current_period,
            "comparative_period_end_date": comparative_period or detected_prior_period,
            "document_type": "CY_DRAFT_FS",
        }
        detected_framework = framework or hints.get("framework")
        if detected_framework:
            metadata["framework"] = detected_framework
        financial = {
            "metadata": metadata,
            "prior_data": {"balance_sheet": {}, "income_statement": {}, "final_trial_balance": {}},
            "current_data": {
                "preliminary_trial_balance": {},
                "balance_sheet": {}, "income_statement": {}, "cash_flow_statement": {},
                "equity_statement": {},
                "footnotes": {"ar_aging": {}, "ppe_sched": {}, "debt_maturity": {}},
            },
        }
        financial["prior_data"]["final_trial_balance"] = _consolidate_trial_balances(trial_balances.get("prior", []))
        financial["current_data"]["preliminary_trial_balance"] = _consolidate_trial_balances(trial_balances.get("current", []))

        report = {"mapped": [], "unmapped": [], "ambiguous": [], "duplicates": [], "ignored": [r for r in source_routes if r["channel"] == "ignored"], "parse_issues": parse_issues, "quality_issues": [], "validation_errors": [], "unit_issues": [], "source_routes": source_routes, "source_files": [str(p) for p in routed_paths], "consolidation": {}}

        # Text-quality audit applies to financial inputs and entity-authored
        # disclosures/MD&A/board notes. Authoritative external standards and
        # bank directives are reference corpus, not client-authored text, so
        # flagging their typography would create noisy false positives.
        for block in text_corpus:
            src_lower = str(block.get("source_file", "")).lower().replace("\\", "/")
            if "/us_gaap/" in src_lower or "/credit_risk_directives/" in src_lower:
                continue
            for issue in lint_text(block.get("text", "")):
                report["quality_issues"].append({"source_file": block.get("source_file"), "location": block.get("location"), **issue})

        # Collect mapped values by entity first. This preserves single-entity behavior
        # while allowing additive consolidation when multiple legal entities are supplied.
        entity_store: dict[str, dict[str, dict[str, float]]] = {}
        detected_entities: set[str] = set()

        for item in items:
            res = self.mapper.map_label(item.label, item.context)
            if res.status != "MAPPED":
                bucket = "ambiguous" if res.status == "AMBIGUOUS" else "unmapped"
                report[bucket].append({**asdict(item), **res.to_dict()})
                continue
            canonical = res.canonical
            if item.source_currency and metadata.get("currency") and item.source_currency != metadata.get("currency"):
                report["unit_issues"].append({"source_file": item.source_file, "source_location": item.source_location, "original_label": item.label, "issue": "CURRENCY_MISMATCH", "source_currency": item.source_currency, "target_currency": metadata.get("currency"), "action": "NO_FX_CONVERSION"})
            scaled_value = convert_scale(item.value, item.source_scale, metadata.get("scale"))
            value, sign_rule = normalize_sign(item.context, canonical, scaled_value)
            if item.context in {"ar_aging", "ppe_sched", "debt_maturity"}:
                target_key = f"current_data.footnotes.{item.context}"
            elif item.context in {"cash_flow_statement", "equity_statement"}:
                target_key = f"current_data.{item.context}"
            else:
                target_key = f"{'prior_data' if item.period_kind == 'prior' else 'current_data'}.{item.context}"

            entity = item.entity_name or metadata.get("entity_name") or metadata.get("client_name") or "__UNSPECIFIED_ENTITY__"
            detected_entities.add(entity)
            entity_target = entity_store.setdefault(target_key, {}).setdefault(entity, {})
            if canonical in entity_target and abs(float(entity_target[canonical]) - float(value)) > 1e-9:
                report["duplicates"].append({**asdict(item), "canonical": canonical, "existing_value": entity_target[canonical], "new_value": value, "action": "KEPT_FIRST_WITHIN_ENTITY"})
            else:
                entity_target.setdefault(canonical, value)
            report["mapped"].append({**asdict(item), "canonical": canonical, "normalized_value": value, "method": res.method, "confidence": res.confidence, "sign_rule": sign_rule, "scale_conversion": f"{item.source_scale or metadata.get('scale')}->{metadata.get('scale')}"})

        def resolve_target(key: str) -> dict:
            parts = key.split(".")
            cur: dict = financial
            for part in parts:
                cur = cur[part]
            return cur

        multi_entity = len(detected_entities) > 1
        for target_key, by_entity in entity_store.items():
            target = resolve_target(target_key)
            if len(by_entity) == 1:
                target.update(next(iter(by_entity.values())))
                continue
            canonical_names = sorted({c for values in by_entity.values() for c in values})
            for canonical in canonical_names:
                target[canonical] = sum(float(values.get(canonical, 0.0)) for values in by_entity.values())

        report["consolidation"] = {
            "mode": "MULTI_ENTITY_ADDITIVE" if multi_entity else "SINGLE_ENTITY",
            "entities_detected": sorted(detected_entities),
            "entity_count": len(detected_entities),
            "intercompany_eliminations_applied": any(
                any(token in norm(e) for token in ("elimination", "consolidation adjustment", "intercompany adjustment"))
                for e in detected_entities
            ),
            "method": "Canonical fields are normalized per entity, then additive fields are summed across entities. Explicit elimination/adjustment entities are included at their supplied sign; Layer 2 does not invent intercompany eliminations.",
        }
        if multi_entity and not report["consolidation"]["intercompany_eliminations_applied"]:
            report["quality_issues"].append({
                "issue_type": "CONSOLIDATION",
                "message": "MULTI_ENTITY_INPUT_WITHOUT_EXPLICIT_INTERCOMPANY_ELIMINATIONS",
                "severity": "MEDIUM",
                "action": "ADDITIVE_CONSOLIDATION_ONLY_NO_ELIMINATIONS_INVENTED",
            })

        # Optional fields can stay absent; do not fabricate required values.
        for period_key in ("prior_data", "current_data"):
            for ctx in ("balance_sheet", "income_statement"):
                missing = sorted(REQUIRED[ctx] - set(financial[period_key][ctx]))
                if missing:
                    report["validation_errors"].append({"path": f"{period_key}.{ctx}", "missing_required_fields": missing})
        missing_cf = sorted(REQUIRED["cash_flow_statement"] - set(financial["current_data"]["cash_flow_statement"]))
        if missing_cf:
            report["validation_errors"].append({"path": "current_data.cash_flow_statement", "missing_required_fields": missing_cf})

        planning = self._build_planning(planning_rows, metadata, hints, report)

        # JSON-schema validation against Layer 3 contracts.
        schema_dir = Path(__file__).resolve().parents[3] / "templates"
        for contract, obj, file in [
            ("FinancialStatementsIngestionSchema", financial, "Input ingestion template.json"),
            ("AnnualOperatingBudget", planning.get("annual_operating_budget"), "aob_schema.json"),
            ("RollingForecast4Q", planning.get("rolling_forecast_4q"), "4q_schema.json"),
            ("RollingForecast8Q", planning.get("rolling_forecast_8q"), "8q_schema.json"),
            ("OperationalDrivers", planning.get("operational_drivers"), "operational_drivers_schema.json"),
        ]:
            if obj is not None:
                for err in validate_schema(obj, schema_dir / file):
                    report["validation_errors"].append({"contract": contract, **err})

        required_metadata = ("client_name", "period", "currency", "scale")
        for field in required_metadata:
            if not metadata.get(field):
                report["validation_errors"].append({
                    "contract": "Metadata",
                    "path": f"metadata.{field}",
                    "message": f"{field.upper()}_NOT_DETECTED",
                })

        if not metadata.get("comparative_period_end_date"):
            report["quality_issues"].append({
                "issue_type": "METADATA",
                "field": "comparative_period_end_date",
                "message": "COMPARATIVE_PERIOD_NOT_DETECTED",
                "severity": "MEDIUM",
            })

        report["summary"] = {
            "files_discovered": len(paths), "files_processed": len(routed_paths), "files_ignored": len(paths) - len(routed_paths), "financial_records_seen": len(items),
            "mapped_count": len(report["mapped"]), "unmapped_count": len(report["unmapped"]),
            "ambiguous_count": len(report["ambiguous"]), "duplicate_count": len(report["duplicates"]),
            "parse_issue_count": len(report["parse_issues"]), "quality_issue_count": len(report["quality_issues"]), "unit_issue_count": len(report["unit_issues"]), "validation_error_count": len(report["validation_errors"]),
            "financial_contract_ready": not any(e.get("contract") in (None, "FinancialStatementsIngestionSchema") for e in report["validation_errors"]),
            "planning_contracts_ready": not any(e.get("contract") in {"AnnualOperatingBudget", "RollingForecast4Q", "RollingForecast8Q", "OperationalDrivers"} for e in report["validation_errors"]),
            "entity_count": len(detected_entities),
            "consolidation_mode": report["consolidation"].get("mode"),
        }
        source_trace = [{k: row.get(k) for k in ("source_file", "source_location", "original_label", "label", "canonical", "normalized_value", "method", "confidence", "sign_rule", "source_currency", "source_scale", "scale_conversion") if row.get(k) is not None} for row in report["mapped"]]
        qualitative_corpus = [b for b in text_corpus if "qualitative_corpus" in str(b.get("source_file", "")).lower()]
        return {"financial_statements": financial, **planning, "normalization_report": report, "source_trace": source_trace, "text_corpus": text_corpus, "qualitative_corpus": qualitative_corpus}

    def _parse_excel(self, path: Path, items: list[Item], trial_balances: dict, planning_rows: list[dict], text_corpus: list[dict], hints: dict):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            lines = [" | ".join(str(v) for v in r if v not in (None, "")) for r in rows if any(v not in (None, "") for v in r)]
            whole = "\n".join(lines[:30])
            local_units = infer_unit_metadata(whole)
            extracted = extract_metadata_text(whole)
            for k, v in extracted.items():
                if k == "period":
                    pk = infer_period_kind(path, whole)
                    hints.setdefault("comparative_period" if pk == "prior" else "current_period", v)
                else:
                    hints.setdefault(k, v)
            if lines:
                text_corpus.append({"source_file": str(path), "location": ws.title, "text": whole})

            sheet_n = norm(ws.title)
            # Layer 1 planning_inputs.xlsx structure.
            if sheet_n in {"aob", "4q forecast", "8q forecast", "operational drivers"}:
                if not rows:
                    continue
                headers = [norm(x) for x in rows[0]]
                col = {h: i for i, h in enumerate(headers) if h}
                for ridx, row in enumerate(rows[1:], start=2):
                    if not any(v not in (None, "") for v in row):
                        continue
                    def get(name):
                        i = col.get(name); return row[i] if i is not None and i < len(row) else None
                    company_value = get("company")
                    if company_value:
                        hints.setdefault("client_name", str(company_value).strip())
                    unit_value = get("unit")
                    if unit_value:
                        for uk, uv in infer_unit_metadata(str(unit_value)).items():
                            if uv:
                                hints.setdefault(uk, uv)
                    if sheet_n == "operational drivers":
                        planning_rows.append({"kind": "operational_driver", "company": company_value, "period": get("period"), "label": get("driver") or get("driver name"), "value": parse_number(get("value")), "unit": get("unit"), "source": f"{path}:{ws.title}!R{ridx}"})
                    else:
                        planning_rows.append({"kind": sheet_n.replace(" ", "_"), "company": company_value, "period": get("period"), "label": get("metric"), "value": parse_number(get("value") if "value" in col else get("amount")), "unit": get("unit"), "source": f"{path}:{ws.title}!R{ridx}"})
                continue

            title = " ".join(str(rows[i][0]) for i in range(min(4, len(rows))) if rows[i] and rows[i][0] not in (None, ""))
            context = infer_context(path, ws.title, title)
            if context == "trial_balance":
                tb = self._parse_trial_balance(rows, path, ws.title, infer_period_kind(path, title), hints)
                tb["metadata"]["entity_name"] = infer_entity(path, whole) or tb["metadata"].get("entity_name")
                trial_balances["prior" if infer_period_kind(path, title) == "prior" else "current"].append(tb)
                continue
            if context not in FINANCIAL_CONTEXTS:
                continue
            period_kind = infer_period_kind(path, title)
            # Find header row; then take first text label + first numeric cell to the right.
            start = 0
            for i, row in enumerate(rows[:15]):
                vals = {norm(v) for v in row if isinstance(v, str)}
                if vals & {"line item", "bucket", "maturity", "field"}:
                    start = i + 1; break
            for ridx, row in enumerate(rows[start:], start=start + 1):
                label = next((str(v).strip() for v in row if isinstance(v, str) and v.strip()), None)
                if not label:
                    continue
                nums = [parse_number(v) for v in row[1:]]
                value = next((x for x in nums if x is not None), None)
                if value is None:
                    continue
                if norm(label) in IGNORED_SUPPLEMENTARY:
                    continue
                items.append(Item(label, value, context, period_kind, str(path), f"{ws.title}!R{ridx}", f"{label} | {value}", local_units.get("currency"), local_units.get("scale"), infer_entity(path, whole)))

    def _parse_trial_balance(self, rows, path: Path, sheet: str, period_kind: str, hints: dict) -> dict:
        header_i = None
        for i, row in enumerate(rows[:15]):
            if any("account code" in norm(v) for v in row if isinstance(v, str)):
                header_i = i; break
        accounts = []
        if header_i is None:
            return {"metadata": {}, "accounts": accounts}
        headers = [norm(x) for x in rows[header_i]]
        col = {h: i for i, h in enumerate(headers) if h}
        def find_col(prefix):
            for h, i in col.items():
                if prefix in h: return i
            return None
        idx = {"code": find_col("account code"), "name": find_col("account name"), "fsli": find_col("standardized fsli"), "type": find_col("account type"), "debit": find_col("debit"), "credit": find_col("credit"), "ending": find_col("ending balance")}
        for row in rows[header_i+1:]:
            if idx["code"] is None or idx["code"] >= len(row) or row[idx["code"]] in (None, ""):
                continue
            def val(k):
                i = idx[k]; return row[i] if i is not None and i < len(row) else None
            accounts.append({
                "account_code": str(val("code")), "account_name": str(val("name") or ""),
                "standardized_fsli": str(val("fsli") or ""), "account_type": str(val("type") or ""),
                "debit_amount": parse_number(val("debit")) or 0.0, "credit_amount": parse_number(val("credit")) or 0.0,
                "ending_balance": parse_number(val("ending")) or 0.0,
            })
        return {"metadata": {"entity_name": hints.get("client_name"), "period_type": "PRIOR_AUDITED" if period_kind == "prior" else "CURRENT_PRELIMINARY", "period_end_date": hints.get("period"), "currency": hints.get("currency"), "scale": hints.get("scale"), "source_file_name": path.name}, "accounts": accounts}

    def _parse_csv(self, path: Path, items: list[Item], text_corpus: list[dict]):
        context = infer_context(path)
        if context not in FINANCIAL_CONTEXTS:
            return
        with path.open(newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        text_corpus.append({"source_file": str(path), "location": "csv", "text": "\n".join(" | ".join(r) for r in rows[:100])})
        for i, row in enumerate(rows, 1):
            if not row: continue
            label = row[0].strip()
            value = next((parse_number(v) for v in row[1:] if parse_number(v) is not None), None)
            if label and value is not None:
                items.append(Item(label, value, context, infer_period_kind(path), str(path), f"row:{i}", source_currency=None, source_scale=None, entity_name=infer_entity(path)))

    def _parse_json(self, path: Path, items: list[Item], planning_rows: list[dict], text_corpus: list[dict]):
        obj = json.loads(path.read_text(encoding="utf-8"))
        text_corpus.append({"source_file": str(path), "location": "json", "text": json.dumps(obj, ensure_ascii=False)[:20000]})
        # Generic recursive label/value extraction when a JSON object is not already a Layer 3 payload.
        context = infer_context(path)
        if context in FINANCIAL_CONTEXTS and isinstance(obj, dict):
            def walk(d):
                for k, v in d.items():
                    if isinstance(v, (int, float)):
                        items.append(Item(k, float(v), context, infer_period_kind(path), str(path), f"json:{k}", source_currency=None, source_scale=None, entity_name=infer_entity(path)))
                    elif isinstance(v, dict): walk(v)
            walk(obj)

    def _parse_jsonl_text(self, path: Path, text_corpus: list[dict]):
        for i, line in enumerate(path.read_text(encoding="utf-8", errors="ignore").splitlines(), 1):
            if not line.strip(): continue
            try: obj = json.loads(line); txt = obj.get("text") or obj.get("content") or line
            except Exception: txt = line
            text_corpus.append({"source_file": str(path), "location": f"line:{i}", "text": str(txt)})

    def _parse_pdf(self, path: Path, items: list[Item], text_corpus: list[dict], hints: dict):
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
        text_corpus.append({"source_file": str(path), "location": "pdf", "text": text})
        for k, v in extract_metadata_text(text).items(): hints.setdefault(k, v)
        context = infer_context(path, title=text[:500])
        local_units = infer_unit_metadata(text[:2000])
        if context not in FINANCIAL_CONTEXTS: return
        for i, line in enumerate(text.splitlines(), 1):
            m = re.match(r"^(.+?)\s+([\(\)-]?[₹$€£]?[\d,]+(?:\.\d+)?\)?)\s*$", line.strip())
            if m:
                val = parse_number(m.group(2))
                if val is not None:
                    items.append(Item(m.group(1).strip(), val, context, infer_period_kind(path, text[:500]), str(path), f"page-text-line:{i}", source_currency=local_units.get("currency"), source_scale=local_units.get("scale"), entity_name=infer_entity(path, text[:2000])))

    def _parse_docx(self, path: Path, text_corpus: list[dict]):
        with zipfile.ZipFile(path) as z:
            xml = z.read("word/document.xml")
        root = ET.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paras = []
        for p in root.findall(".//w:p", ns):
            text = "".join(t.text or "" for t in p.findall(".//w:t", ns))
            if text.strip(): paras.append(text.strip())
        text_corpus.append({"source_file": str(path), "location": "docx", "text": "\n".join(paras)})

    def _parse_text(self, path: Path, text_corpus: list[dict]):
        text_corpus.append({"source_file": str(path), "location": "text", "text": path.read_text(encoding="utf-8", errors="ignore")})

    def _build_planning(self, rows: list[dict], metadata: dict, hints: dict, report: dict) -> dict:
        company = metadata.get("client_name")
        currency = metadata.get("currency")
        scale = metadata.get("scale")
        aob_lines, aob_drivers, op_drivers = [], [], []
        q4: list[dict] = []; q8: list[dict] = []
        for r in rows:
            if r.get("value") is None or not r.get("label"): continue
            label_n = norm(r["label"]); period = str(r.get("period") or "")
            if r["kind"] == "aob":
                drv = DRIVER_MAP.get(label_n)
                if drv and drv[1]:
                    aob_drivers.append({"driver_name": str(r["label"]), "driver_code": drv[1], "value": float(r["value"]), "unit": str(r.get("unit") or "")})
                    continue
                m = self.mapper.map_label(str(r["label"]), "budget_metric")
                if m.status == "MAPPED":
                    aob_lines.append({"metric": str(r["label"]), "canonical_code": m.canonical, "category": BUDGET_CATEGORY[m.canonical], "amount": float(r["value"])})
                else:
                    report["unmapped"].append({"source": r.get("source"), "label": r["label"], "context": "budget_metric", **m.to_dict()})
            elif r["kind"] == "operational_driver":
                drv = DRIVER_MAP.get(label_n)
                if not drv:
                    m = self.mapper.map_label(str(r["label"]), "operational_driver")
                    driver_type = m.canonical if m.status == "MAPPED" else None
                else:
                    driver_type = drv[0]
                if driver_type:
                    op_drivers.append({"period": period, "driver_type": driver_type, "driver_name": str(r["label"]), "value": float(r["value"]), "unit": str(r.get("unit") or ""), "granularity": "QUARTERLY" if "-Q" in period else "ANNUAL"})
            elif r["kind"] in {"4q_forecast", "8q_forecast"}:
                field = FORECAST_MAP.get(label_n)
                if not field:
                    mi = self.mapper.map_label(str(r["label"]), "income_statement")
                    if mi.status == "MAPPED":
                        field = {"revenue": "revenue", "cogs": "cogs", "total_operating_expenses": "opex", "operating_income": "operating_income"}.get(mi.canonical)
                if not field: continue
                target = q4 if r["kind"] == "4q_forecast" else q8
                rec = next((x for x in target if x["period"] == period), None)
                if rec is None:
                    qm = re.search(r"Q([1-4])$", period); fy = period.split("-")[0] if period else ""
                    rec = {"period": period, "quarter": int(qm.group(1)) if qm else None}
                    if target is q8: rec["fiscal_year"] = fy
                    target.append(rec)
                rec[field] = float(r["value"])

        out = {"annual_operating_budget": None, "rolling_forecast_4q": None, "rolling_forecast_8q": None, "operational_drivers": None}
        if aob_lines or aob_drivers:
            fy = next((str(r.get("period")) for r in rows if r["kind"] == "aob" and r.get("period")), hints.get("fiscal_year"))
            out["annual_operating_budget"] = {"company": company, "fiscal_year": fy, "currency": currency, "scale": scale, "line_items": aob_lines, "drivers": aob_drivers}
        if q4:
            q4.sort(key=lambda x: x["period"])
            base = q4[0]["period"].split("-")[0] if q4 else ""
            out["rolling_forecast_4q"] = {"company": company, "base_fiscal_year": base, "currency": currency, "scale": scale, "quarterly_projections": q4}
        if q8:
            q8.sort(key=lambda x: x["period"])
            out["rolling_forecast_8q"] = {"company": company, "start_period": q8[0]["period"], "end_period": q8[-1]["period"], "currency": currency, "scale": scale, "quarterly_projections": q8}
        if op_drivers:
            out["operational_drivers"] = {"company": company, "drivers": op_drivers}
        return out


def validate_schema(instance: Any, schema_path: Path) -> list[dict]:
    try:
        import jsonschema
    except ImportError:
        return [{"path": "$", "message": "jsonschema is not installed"}]
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    validator_cls = jsonschema.validators.validator_for(schema)
    validator_cls.check_schema(schema)
    validator = validator_cls(schema, format_checker=jsonschema.FormatChecker())
    errors = []
    for e in sorted(validator.iter_errors(instance), key=lambda x: list(x.absolute_path)):
        p = "$" + "".join(f"[{x}]" if isinstance(x, int) else f".{x}" for x in e.absolute_path)
        errors.append({"path": p, "message": e.message})
    return errors


def write_outputs(result: dict, output_dir: str | Path) -> dict[str, str]:
    out = Path(output_dir); out.mkdir(parents=True, exist_ok=True)
    names = {
        "financial_statements": "financial_statements.json",
        "annual_operating_budget": "annual_operating_budget.json",
        "rolling_forecast_4q": "rolling_forecast_4q.json",
        "rolling_forecast_8q": "rolling_forecast_8q.json",
        "operational_drivers": "operational_drivers.json",
        "normalization_report": "normalization_report.json",
        "source_trace": "source_trace.json",
        "text_corpus": "extracted_text_corpus.json",
        "qualitative_corpus": "qualitative_corpus.json",
    }
    written = {}
    for key, fname in names.items():
        obj = result.get(key)
        if obj is None: continue
        p = out / fname
        p.write_text(json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8")
        written[key] = str(p)
    return written
