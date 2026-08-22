import io
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelModelGenerator:
    """Generate the WP-514 supporting/reconciled audit workbook from upstream results."""

    @classmethod
    def generate_reconciled_workbook(
        cls,
        engagement_info: Dict[str, Any],
        report_data: Dict[str, Any],
        waiver_records: Optional[List[Dict[str, Any]]] = None,
    ) -> bytes:
        engagement = dict(report_data.get("engagement") or {})
        for k, v in (engagement_info or {}).items():
            if v not in (None, ""):
                engagement[k] = v

        client = str(engagement.get("client_name") or "Client not provided")
        period = str(engagement.get("period") or "Period not provided")
        currency = str(engagement.get("currency") or "")
        scale = str(engagement.get("scale") or "")
        framework = str(engagement.get("framework") or "Not provided")
        review_stage = str(engagement.get("review_stage") or "Not provided")
        procedures = report_data.get("procedures") or []
        analytics = report_data.get("analytics") or {}
        findings = report_data.get("findings") or []
        conclusion = report_data.get("conclusion") or {}
        waivers = waiver_records or []

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        navy = "1F4E78"
        light = "F4F7FA"
        border_color = "C8D2DC"
        green = "E2F0D9"
        amber = "FCE4D6"
        red = "F4CCCC"
        title_font = Font(name="Calibri", size=16, bold=True, color=navy)
        subtitle_font = Font(name="Calibri", size=11, italic=True, color="475569")
        header_fill = PatternFill("solid", fgColor=navy)
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        bold = Font(name="Calibri", size=10, bold=True)
        regular = Font(name="Calibri", size=10)
        thin = Border(left=Side(style="thin", color=border_color), right=Side(style="thin", color=border_color), top=Side(style="thin", color=border_color), bottom=Side(style="thin", color=border_color))

        # Values are already expressed in the upstream scale; do not rescale here.
        number_fmt = '#,##0.00;[Red](#,##0.00);-'
        pct_fmt = '0.0%;[Red](0.0%);-'

        def style_header(ws, row, headers):
            for col, text in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=text)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = thin

        def status_color(status: str):
            s = str(status or "").upper()
            if s in {"PASS", "CLEARED", "RESOLVED"}:
                return green
            if s in {"WARNING", "FLAGGED", "REVIEW REQUIRED", "WAIVED", "YES"}:
                return amber
            if s in {"FAIL", "REJECTED", "CRITICAL", "OPEN"}:
                return red
            return light

        def finalize(ws, freeze=None, widths=None):
            ws.sheet_view.showGridLines = False
            if freeze:
                ws.freeze_panes = freeze
            if widths:
                for col, width in widths.items():
                    ws.column_dimensions[col].width = width
            ws.auto_filter.ref = ws.dimensions if ws.max_row > 1 else None

        # 1. Executive Summary
        ws = wb.create_sheet("Executive Summary")
        ws.merge_cells("A2:F2")
        ws["A2"] = f"{client} - WP-514 AUDIT WORKPAPER"
        ws["A2"].font = title_font
        ws.merge_cells("A3:F3")
        ws["A3"] = "Financial Statement Review & Audit Assurance Summary"
        ws["A3"].font = subtitle_font
        meta = [
            ("Client / Entity", client), ("Reporting Period", period), ("Currency", currency),
            ("Scale", scale), ("Framework", framework), ("Review Stage", review_stage),
            ("Overall Status", conclusion.get("overall_status", "")),
            ("Procedures Executed", conclusion.get("total_procedures_run", len(procedures))),
            ("Procedures Passed", conclusion.get("procedures_passed", sum(1 for p in procedures if p.get("status") == "PASS"))),
            ("Findings", len(findings)), ("Waivers / AJEs", len(waivers)),
        ]
        r = 5
        for label, value in meta:
            ws.cell(r, 1, label).font = bold
            ws.cell(r, 2, value).font = regular
            r += 1
        r += 1
        ws.cell(r, 1, "Audit Conclusion").font = bold
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+4, end_column=6)
        ws.cell(r+1, 1, conclusion.get("text") or "No conclusion text supplied.")
        ws.cell(r+1, 1).alignment = Alignment(wrap_text=True, vertical="top")
        finalize(ws, widths={"A": 25, "B": 28, "C": 16, "D": 16, "E": 18, "F": 18})

        # 2. Review checklist - mirrors sample WP-514 checklist but uses all upstream procedures.
        ws = wb.create_sheet("WP-514 Review Checklist")
        ws.merge_cells("A1:G1")
        ws["A1"] = "FINANCIAL STATEMENT REVIEW & QUALITY CONTROL CHECKLIST"
        ws["A1"].font = title_font
        headers = ["#", "Procedure Category", "Specific Review Control Test", "Ref", "Status", "Issue", "Resolution"]
        style_header(ws, 3, headers)
        for i, p in enumerate(procedures, 4):
            vals = [p.get("step"), p.get("category"), p.get("procedure"), p.get("reference"), p.get("status"), p.get("issue"), p.get("resolution")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(i, c, v)
                cell.border = thin
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(i, 5).fill = PatternFill("solid", fgColor=status_color(p.get("status")))
            ws.cell(i, 5).font = bold
        finalize(ws, freeze="A4", widths={"A": 6, "B": 27, "C": 55, "D": 14, "E": 14, "F": 35, "G": 35})

        # 3. Financial analytics - historical IS + BS only; no forecasts/BvA planning.
        ws = wb.create_sheet("Financial Analytics")
        row = 1
        for section_name, records in [("INCOME STATEMENT YOY ANALYTICS", analytics.get("income_statement") or []), ("BALANCE SHEET YOY ANALYTICS", analytics.get("balance_sheet") or [])]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            ws.cell(row, 1, section_name).font = title_font
            row += 2
            style_header(ws, row, ["Financial Line Item", "Prior Period", "Current Period", "Variance", "Variance (%)", "Threshold Status", "Commentary"])
            row += 1
            for item in records:
                vals = [item.get("line_item"), item.get("prior_period"), item.get("current_period"), item.get("variance"), (item.get("variance_pct") / 100.0 if isinstance(item.get("variance_pct"), (int, float)) else item.get("variance_pct")), item.get("threshold_status"), item.get("commentary")]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row, c, v)
                    cell.border = thin
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                for c in (2, 3, 4):
                    ws.cell(row, c).number_format = number_fmt
                ws.cell(row, 5).number_format = pct_fmt
                ws.cell(row, 6).fill = PatternFill("solid", fgColor=status_color(item.get("threshold_status")))
                row += 1
            row += 2
        finalize(ws, freeze="A4", widths={"A": 30, "B": 17, "C": 17, "D": 17, "E": 15, "F": 17, "G": 55})

        # 4. Ratio analytics
        ws = wb.create_sheet("Ratio Analytics")
        ws.merge_cells("A1:H1")
        ws["A1"] = "FINANCIAL RATIO ANALYTICS"
        ws["A1"].font = title_font
        style_header(ws, 3, ["Category", "Metric", "Formula", "Prior Period", "Current Period", "Benchmark", "Status", "Assessment"])
        for i, item in enumerate(analytics.get("ratios") or [], 4):
            vals = [item.get("category"), item.get("name"), item.get("formula"), item.get("prior_period"), item.get("current_period"), item.get("benchmark"), item.get("status"), item.get("assessment")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(i, c, v)
                cell.border = thin
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(i, 7).fill = PatternFill("solid", fgColor=status_color(item.get("status")))
            ws.cell(i, 7).font = bold
        finalize(ws, freeze="A4", widths={"A": 18, "B": 24, "C": 42, "D": 15, "E": 15, "F": 24, "G": 14, "H": 48})

        # 5. Findings and exceptions
        ws = wb.create_sheet("Findings & Exceptions")
        ws.merge_cells("A1:K1")
        ws["A1"] = "AUDIT FINDINGS & EXCEPTIONS"
        ws["A1"].font = title_font
        style_header(ws, 3, ["ID", "Category", "Severity", "Description", "Expected", "Actual", "Difference", "Confidence", "Status", "Recommended Action", "Evidence"])
        if findings:
            for i, f in enumerate(findings, 4):
                vals = [f.get("id"), f.get("category"), f.get("severity"), f.get("description"), f.get("expected"), f.get("actual"), f.get("difference"), f.get("confidence"), f.get("status"), f.get("recommended_action"), "; ".join(str(x) for x in (f.get("evidence") or []))]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(i, c, v)
                    cell.border = thin
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.cell(i, 3).fill = PatternFill("solid", fgColor=status_color(f.get("severity")))
                ws.cell(i, 9).fill = PatternFill("solid", fgColor=status_color(f.get("status")))
        else:
            ws.cell(4, 1, "No audit findings were reported by the upstream audit engine.")
        finalize(ws, freeze="A4", widths={"A": 15, "B": 30, "C": 12, "D": 50, "E": 14, "F": 14, "G": 14, "H": 13, "I": 13, "J": 48, "K": 48})

        # 6. Waivers & AJEs - no fabricated sign-off names or values.
        ws = wb.create_sheet("Waivers & AJEs")
        ws.merge_cells("A1:H1")
        ws["A1"] = "AUDIT WAIVER LEDGER & AJE LOG"
        ws["A1"].font = title_font
        style_header(ws, 3, ["Rule ID", "Decision", "Submitted Value", "Expected Value", "Variance", "Justification / Notes", "Resolved By", "Timestamp"])
        if waivers:
            for i, w in enumerate(waivers, 4):
                submitted = w.get("submitted_value")
                expected = w.get("expected_value")
                variance = (submitted - expected) if isinstance(submitted, (int, float)) and isinstance(expected, (int, float)) else None
                vals = [w.get("rule_id"), w.get("user_decision"), submitted, expected, variance, w.get("justification_notes"), w.get("resolved_by"), str(w.get("resolved_at") or "")]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(i, c, v)
                    cell.border = thin
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                for c in (3, 4, 5):
                    ws.cell(i, c).number_format = number_fmt
        else:
            ws.cell(4, 1, "No waiver or AJE records were supplied for this engagement.")
        finalize(ws, freeze="A4", widths={"A": 16, "B": 16, "C": 18, "D": 18, "E": 18, "F": 45, "G": 22, "H": 24})

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
