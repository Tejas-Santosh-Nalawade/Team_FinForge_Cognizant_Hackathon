"""
Excel Dataset Loader for MathEngine.
Parses financial statement Excel files from data folders (e.g., Data/Error_data, Data/True_data)
and constructs validated FinancialStatementsIngestionSchema models.
"""

import json
from pathlib import Path
from typing import Any, Dict, Union
import openpyxl

from math_engine.schemas import (
    FinancialStatementsIngestionSchema,
    Metadata,
    PriorData,
    CurrentData,
    BalanceSheetValues,
    IncomeStatementValues,
    CashFlowValues,
    StockholdersEquity,
    Footnotes,
    AccountsReceivableAging,
    PPESchedule,
    DebtMaturities,
)


def parse_excel_key_values(filepath: Path) -> Dict[str, float]:
    """Extract key-value pairs from standard two-column financial statement Excel sheets."""
    if not filepath.exists():
        return {}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    data = {}
    for row in sheet.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        key, val = row[0], row[1]
        if key is not None and val is not None and isinstance(val, (int, float)):
            data[str(key).strip()] = float(val)
    return data


def parse_trial_balance(filepath: Path) -> Dict[str, Any]:
    """Parse trial balance excel into dictionary structure."""
    if not filepath.exists():
        return {}

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    rows = []
    header_found = False
    for row in sheet.iter_rows(values_only=True):
        if not row or not any(row):
            continue
        first_col = str(row[0]).strip() if row[0] is not None else ""
        if first_col == "Account Code":
            header_found = True
            continue
        if header_found and len(row) >= 7:
            rows.append({
                "account_code": str(row[0]),
                "account_name": str(row[1]),
                "fsli": str(row[2]),
                "account_type": str(row[3]),
                "debit": float(row[4]) if row[4] is not None else 0.0,
                "credit": float(row[5]) if row[5] is not None else 0.0,
                "ending_balance": float(row[6]) if row[6] is not None else 0.0,
            })
    return {"accounts": rows}


def map_balance_sheet(data: Dict[str, float]) -> BalanceSheetValues:
    return BalanceSheetValues(
        cash_and_cash_equivalents=data.get("Cash & Cash Equivalents", 0.0),
        accounts_receivable_net=data.get("Accounts Receivable, net", 0.0),
        inventory=data.get("Inventory", 0.0),
        prepaid_expenses=data.get("Prepaid Expenses & Other Current Assets", data.get("Prepaid Expenses", 0.0)),
        total_current_assets=data.get("Total Current Assets", 0.0),
        ppe_net=data.get("Property, Plant & Equipment, net", 0.0),
        intangible_assets=data.get("Intangible Assets", 0.0),
        total_non_current_assets=data.get("Property, Plant & Equipment, net", 0.0) + data.get("Intangible Assets", 0.0),
        total_assets=data.get("Total Assets", 0.0),
        accounts_payable=data.get("Accounts Payable", 0.0),
        accrued_expenses=data.get("Accrued Liabilities", data.get("Accrued Expenses", 0.0)),
        short_term_debt=data.get("Short-term Debt", 0.0),
        current_portion_of_lt_debt=data.get("Current Portion of Long-Term Debt", 0.0),
        total_current_liabilities=data.get("Total Current Liabilities", 0.0),
        long_term_debt=data.get("Long-term Debt", 0.0),
        total_non_current_liabilities=data.get("Long-term Debt", 0.0),
        total_liabilities=data.get("Total Liabilities", 0.0),
        common_stock=data.get("Common Stock", 0.0),
        additional_paid_in_capital=data.get("Additional Paid-in Capital", 0.0),
        retained_earnings=data.get("Retained Earnings", 0.0),
        total_equity=data.get("Total Equity", 0.0),
    )


def map_income_statement(data: Dict[str, float]) -> IncomeStatementValues:
    return IncomeStatementValues(
        revenue=data.get("Revenue", 0.0),
        cogs=data.get("Cost of Goods Sold", 0.0),
        gross_profit=data.get("Gross Profit", 0.0),
        sga_expense=data.get("Selling, General & Administrative", 0.0),
        rd_expense=data.get("Research & Development", 0.0),
        depreciation_amortization=data.get("Depreciation & Amortization", 0.0),
        total_operating_expenses=data.get("Total Operating Expenses", 0.0),
        operating_income=data.get("Operating Income", 0.0),
        interest_expense=data.get("Interest Expense", 0.0),
        non_operating_income=data.get("Non-Operating Income / (Expense)", 0.0),
        income_tax_expense=data.get("Income Tax Expense", 0.0),
        net_income=data.get("Net Income", 0.0),
    )


def map_cash_flow_statement(data: Dict[str, float]) -> CashFlowValues:
    return CashFlowValues(
        net_income_starting=data.get("Net Income", 0.0),
        depreciation_addback=data.get("Depreciation & Amortization Add-back", 0.0),
        working_capital_changes=data.get("Working Capital Changes", 0.0),
        operating_cash_flow=data.get("Operating Cash Flow", 0.0),
        capital_expenditures=data.get("Capital Expenditures", 0.0),
        investing_cash_flow=data.get("Investing Cash Flow", 0.0),
        debt_repayments_or_borrowings=data.get("Debt Borrowings / (Repayments)", 0.0),
        dividends_paid=data.get("Dividends Paid", 0.0),
        financing_cash_flow=data.get("Financing Cash Flow", 0.0),
        net_cash_change=data.get("Net Cash Change", 0.0),
        beginning_cash=data.get("Beginning Cash", 0.0),
        ending_cash=data.get("Ending Cash", 0.0),
    )


def map_equity_statement(data: Dict[str, float]) -> StockholdersEquity:
    divs = data.get("Dividends Declared", 0.0)
    return StockholdersEquity(
        beginning_retained_earnings=data.get("Beginning Retained Earnings", 0.0),
        net_income=data.get("Net Income", 0.0),
        dividends_declared=abs(divs),
        ending_retained_earnings=data.get("Ending Retained Earnings", 0.0),
    )


def map_footnotes(curr_dir: Path) -> Footnotes:
    footnotes_dir = curr_dir / "footnotes"
    ar_data = parse_excel_key_values(footnotes_dir / "ar_aging.xlsx")
    ppe_data = parse_excel_key_values(footnotes_dir / "ppe_sched.xlsx")
    debt_data = parse_excel_key_values(footnotes_dir / "debt_maturity.xlsx")

    ar_aging = AccountsReceivableAging(
        current=ar_data.get("Current", 0.0),
        days_1_30=ar_data.get("1-30 Days", 0.0),
        days_31_60=ar_data.get("31-60 Days", 0.0),
        days_61_90=ar_data.get("61-90 Days", 0.0),
        days_over_90=ar_data.get("Over 90 Days", 0.0),
        gross_ar=ar_data.get("Gross Accounts Receivable", 0.0),
        allowance_for_credit_losses=abs(ar_data.get("Allowance for Credit Losses", 0.0)),
        net_ar=ar_data.get("Net Accounts Receivable", 0.0),
    ) if ar_data else None

    ppe_sched = PPESchedule(
        gross_ppe=ppe_data.get("Gross PP&E", 0.0),
        accumulated_depreciation=abs(ppe_data.get("Accumulated Depreciation", 0.0)),
        net_ppe=ppe_data.get("Net PP&E", 0.0),
        additions_capex=ppe_data.get("Additions / CapEx", 0.0),
        disposals=abs(ppe_data.get("Disposals", 0.0)),
        depreciation_expense=abs(ppe_data.get("Depreciation Expense", 0.0)),
    ) if ppe_data else None

    debt_maturity = DebtMaturities(
        year_1=debt_data.get("Year 1", 0.0),
        year_2=debt_data.get("Year 2", 0.0),
        year_3=debt_data.get("Year 3", 0.0),
        year_4=debt_data.get("Year 4", 0.0),
        year_5=debt_data.get("Year 5", 0.0),
        thereafter=debt_data.get("Thereafter", 0.0),
        total_debt=debt_data.get("Total Debt", 0.0),
    ) if debt_data else None

    return Footnotes(
        ar_aging=ar_aging,
        ppe_sched=ppe_sched,
        debt_maturity=debt_maturity,
    )


def load_dataset_from_folder(folder_path: Union[str, Path]) -> FinancialStatementsIngestionSchema:
    """
    Load financial dataset from a folder containing prior_data and current_data subdirectories.
    
    Args:
        folder_path: Path to dataset directory (e.g., Data/Error_data or Data/True_data)
        
    Returns:
        Validated FinancialStatementsIngestionSchema instance.
    """
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Dataset directory '{folder}' does not exist.")

    prior_dir = folder / "prior_data"
    curr_dir = folder / "current_data"

    # Prior Period Data
    prior_bs = map_balance_sheet(parse_excel_key_values(prior_dir / "balance_sheet.xlsx"))
    prior_inc = map_income_statement(parse_excel_key_values(prior_dir / "income_statement.xlsx"))
    prior_tb = parse_trial_balance(prior_dir / "final_trial_balance.xlsx")

    prior_data = PriorData(
        balance_sheet=prior_bs,
        income_statement=prior_inc,
        final_trial_balance=prior_tb,
    )

    # Current Period Data
    curr_bs = map_balance_sheet(parse_excel_key_values(curr_dir / "balance_sheet.xlsx"))
    curr_inc = map_income_statement(parse_excel_key_values(curr_dir / "income_statement.xlsx"))
    curr_cf = map_cash_flow_statement(parse_excel_key_values(curr_dir / "cash_flow_statement.xlsx"))
    curr_eq = map_equity_statement(parse_excel_key_values(curr_dir / "equity_statement.xlsx"))
    curr_tb = parse_trial_balance(curr_dir / "preliminary_trial_balance.xlsx")
    footnotes = map_footnotes(curr_dir)

    current_data = CurrentData(
        preliminary_trial_balance=curr_tb,
        balance_sheet=curr_bs,
        income_statement=curr_inc,
        cash_flow_statement=curr_cf,
        equity_statement=curr_eq,
        footnotes=footnotes,
    )

    # Attach AOB and Operational Drivers data if present
    aob_file = folder / "aob.xlsx"
    drivers_file = folder / "operational_drivers.xlsx"

    if aob_file.exists():
        aob_dict = parse_excel_key_values(aob_file)
        setattr(current_data, "aob", aob_dict)
        setattr(current_data, "annual_operating_budget", aob_dict)

    if drivers_file.exists():
        drv_dict = parse_excel_key_values(drivers_file)
        setattr(current_data, "operational_drivers", drv_dict)
        setattr(current_data, "drivers", drv_dict)

    metadata = Metadata(
        client_name="AsterNova Technologies Ltd.",
        period="2026-03-31",
        currency="INR",
        scale="EXACT",
        framework="US GAAP / IFRS",
        review_stage="CY_DRAFT_FS",
    )

    return FinancialStatementsIngestionSchema(
        metadata=metadata,
        prior_data=prior_data,
        current_data=current_data,
    )
